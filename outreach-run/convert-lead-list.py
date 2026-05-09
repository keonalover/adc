import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


HEADERS = [
    "Company",
    "Primary Contact",
    "Email",
    "Phone",
    "City / Market",
    "State",
    "Website",
    "POS",
    "Locations",
    "Lead Source",
    "Temperature",
    "Stage",
    "Estimated Monthly Value",
    "Likely Red Flags / Pain",
    "Personalization Notes",
    "Next Action",
    "Next Action Date",
    "Do Not Contact?",
    "Owner / Internal Notes",
]

uid_counter = 0


def main():
    input_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("../ADC Outreach Lead List Template.xlsx")
    output_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(".")
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(input_path, data_only=True)
    if "Lead List" not in workbook.sheetnames:
        raise SystemExit("Could not find a sheet named 'Lead List'.")

    sheet = workbook["Lead List"]
    headers = [clean(cell.value, keep_na=True) for cell in sheet[1][: len(HEADERS)]]
    if headers != HEADERS:
        raise SystemExit("Lead List headers do not match the ADC template.")

    today = date.today().isoformat()
    leads = []
    warnings = []
    skipped = []
    seen_emails = {}

    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row_number, col).value for col in range(1, len(HEADERS) + 1)]
        record = dict(zip(HEADERS, values))
        company = clean(record["Company"], keep_na=True)
        email = clean(record["Email"])
        do_not_contact = is_yes(record["Do Not Contact?"])

        if not company and not email:
            continue
        if looks_like_example(company, email):
            skipped.append(batch_item(row_number, company, email, "template_row", "Example/template row skipped."))
            warnings.append(warning(row_number, company, email, "skip", "Example/template row skipped."))
            continue
        if do_not_contact:
            skipped.append(batch_item(row_number, company, email, "do_not_contact", "Do Not Contact = Yes."))
            warnings.append(warning(row_number, company, email, "skip", "Do Not Contact = Yes."))
            continue
        if not company:
            warnings.append(warning(row_number, company, email, "warning", "Missing company."))
        if not is_valid_email(email):
            warnings.append(warning(row_number, company, email, "warning", "Missing or invalid email."))
        if email:
            key = email.lower()
            if key in seen_emails:
                skipped.append(batch_item(row_number, company, email, "duplicate_email", f"Duplicate email also appears on row {seen_emails[key]}."))
                warnings.append(warning(row_number, company, email, "skip", f"Duplicate email also appears on row {seen_emails[key]}."))
                continue
            seen_emails[key] = row_number

        lead = {
            "id": uid(),
            "company": company,
            "contact": clean(record["Primary Contact"]),
            "email": email,
            "phone": clean(record["Phone"]),
            "city": clean(record["City / Market"], keep_na=True),
            "state": clean(record["State"], keep_na=True),
            "website": clean(record["Website"], keep_na=True),
            "pos": clean(record["POS"], keep_na=True) or "Unknown",
            "locations": to_number(record["Locations"], 1),
            "source": clean(record["Lead Source"], keep_na=True),
            "temperature": clean(record["Temperature"], keep_na=True) or "Cold",
            "stage": clean(record["Stage"], keep_na=True) or "Research",
            "value": to_number(record["Estimated Monthly Value"], 0),
            "pain": clean(record["Likely Red Flags / Pain"], keep_na=True),
            "personalization": clean(record["Personalization Notes"], keep_na=True),
            "nextAction": clean(record["Next Action"], keep_na=True) or "Send intro email",
            "nextDate": normalize_date(record["Next Action Date"]) or today,
            "nextDueAt": normalize_date(record["Next Action Date"]) or today,
            "doNotContact": False,
            "active": True,
            "replied": False,
            "bounced": False,
            "paused": False,
            "notes": clean(record["Owner / Internal Notes"], keep_na=True),
            "sequenceStep": 0,
            "touches": 0,
            "touchCount": 0,
            "lastTouch": "",
            "lastSentAt": "",
            "lastGmailMessageId": "",
            "threadId": "",
            "updatedAt": int(datetime.now().timestamp() * 1000),
        }
        leads.append(lead)

    base = slug(input_path.stem) or "adc-leads"
    json_path = output_dir / f"{base}-crm-import.json"
    summary_path = output_dir / f"{base}-summary.json"
    draft_path = output_dir / f"{base}-draft-preview.json"
    batch_path = output_dir / f"{base}-batch-review.json"

    drafts = []
    sendable = []
    needs_manual_contact = []
    same_company = {}
    for lead in leads:
        if not is_valid_email(lead["email"]):
            needs_manual_contact.append(batch_item(None, lead["company"], lead["email"], "missing_or_invalid_email", "Needs manual contact form or better email."))
            continue
        same_company.setdefault(lead["company"].lower(), []).append(lead["email"])
        draft = generate_draft(lead)
        subject_line, body = draft.split("\n", 1)
        draft_item = {
            "company": lead["company"],
            "contact": lead["contact"],
            "to": lead["email"],
            "subject": subject_line.replace("Subject: ", ""),
            "body": body.strip(),
            "touch": 0,
            "nextDueAt": lead["nextDueAt"],
        }
        drafts.append(draft_item)
        sendable.append({k: draft_item[k] for k in ["company", "contact", "to", "subject", "touch", "nextDueAt"]})

    same_company_warnings = [
        {"company": company, "emails": emails, "message": "Same company has multiple sendable emails; review before batch send."}
        for company, emails in same_company.items()
        if len(emails) > 1
    ]

    json_path.write_text(json.dumps(leads, ensure_ascii=False, indent=2), encoding="utf-8")
    draft_path.write_text(json.dumps(drafts, ensure_ascii=False, indent=2), encoding="utf-8")
    batch_path.write_text(json.dumps({
        "input": str(input_path.resolve()),
        "createdAt": datetime.now().isoformat(),
        "sendable": sendable,
        "skipped": skipped,
        "needsManualContact": needs_manual_contact,
        "sameCompanyWarnings": same_company_warnings,
        "replyOrBounceStopped": [],
        "approvalRequired": True,
        "sendRule": "No real Gmail sends without explicit approval of this batch.",
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_path.write_text(json.dumps({
        "input": str(input_path.resolve()),
        "createdAt": datetime.now().isoformat(),
        "totalImportable": len(leads),
        "validEmailCount": sum(1 for lead in leads if is_valid_email(lead["email"])),
        "draftPreviewCount": len(drafts),
        "sendableCount": len(sendable),
        "needsManualContactCount": len(needs_manual_contact),
        "skippedCount": len(skipped),
        "sameCompanyWarningCount": len(same_company_warnings),
        "warningCount": len(warnings),
        "warnings": warnings,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({
        "jsonPath": str(json_path.resolve()),
        "summaryPath": str(summary_path.resolve()),
        "draftPreviewPath": str(draft_path.resolve()),
        "batchReviewPath": str(batch_path.resolve()),
        "leads": len(leads),
        "draftPreviews": len(drafts),
        "sendable": len(sendable),
        "warnings": len(warnings),
    }, indent=2))


def clean(value, keep_na=False):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not keep_na and text.lower() in {"n/a", "na", "none", "null"}:
        return ""
    return text


def is_yes(value):
    return clean(value).lower() in {"yes", "y", "true", "1"}


def is_valid_email(email):
    return re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", clean(email)) is not None


def normalize_date(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        return text[:10]
    return ""


def to_number(value, default):
    try:
        if value in (None, ""):
            return default
        return int(value) if float(value).is_integer() else float(value)
    except (TypeError, ValueError):
        return default


def warning(row, company, email, severity, message):
    return {"row": row, "company": company, "email": email, "severity": severity, "message": message}


def batch_item(row, company, email, reason, message):
    return {"row": row, "company": company, "email": email, "reason": reason, "message": message}


def looks_like_example(company, email):
    return bool(re.search(r"example|sample", company or "", re.I) or re.search(r"example\.com$", email or "", re.I))


def uid():
    global uid_counter
    uid_counter += 1
    return f"lead-{int(datetime.now().timestamp() * 1000)}-{uid_counter}"


def slug(value):
    value = re.sub(r"[^a-z0-9]+", "-", value.lower())
    return value.strip("-")


def generate_draft(lead):
    first_name = (lead["contact"] or "there").split(" ")[0]
    personalization = personalization_line(lead["personalization"], lead["company"])
    return (
        f"Subject: quick question for {lead['company']}\n\n"
        f"Hi {first_name},{personalization}\n\n"
        "I'm An, and I'm working on ADC to help F&B owners make better use of the reports they already have: POS, labor, invoices, delivery, inventory, that kind of thing.\n\n"
        "Right now I'm doing the review hands-on for free while I'm getting it off the ground.\n\n"
        "If you'd like to learn more, here's the page:\n"
        "https://adc-consulting.netlify.app/\n\n"
        "Or if you just want to reply with your reports, that works too.\n\n"
        "Best,\nAn"
    )


def personalization_line(note, company):
    text = clean(note, keep_na=True)
    if not text:
        return ""
    lower = text.lower()
    blocked = ["unsure", "probably", "need to", "damn", "target client", "contact form", "would be cool", "skip breakfast"]
    if any(term in lower for term in blocked):
        return ""
    sentence = re.split(r"[.!?]\s", text)[0].strip()
    sentence = re.sub(r"[.!?]+$", "", sentence).strip()
    if not sentence or len(sentence) > 170:
        return ""
    replacements = [
        (r"^also does\s+", f"{company} also does "),
        (r"^also\s+", ""),
        (r"^first\s+founded", "I saw that you first founded"),
        (r"^family also owns", "I noticed your family also owns"),
        (r"^husband\s*&\s*wife duo$", "I saw that the business is run by a husband-and-wife team"),
        (r"^founded by father", "I noticed the business was founded by your father"),
        (r"^used to", "I saw that you used to"),
    ]
    for pattern, replacement in replacements:
        sentence = re.sub(pattern, replacement, sentence, flags=re.I)
    if not re.match(r"^I\s", sentence):
        starts_with_company = company and sentence.lower().startswith(company.lower())
        sentence = f"I noticed {sentence if starts_with_company else sentence[:1].lower() + sentence[1:]}"
    return f"\n\n{sentence}."


if __name__ == "__main__":
    main()
